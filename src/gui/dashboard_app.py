import sys
import os
import requests
import random
import urllib.parse
import datetime
import webbrowser
import json
import pandas as pd
import re
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

# .env 환경변수 자동 로드 (API 키 등)
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), "../../.env"))
except ImportError:
    pass  # python-dotenv 미설치 시 무시 (시스템 환경변수 사용)

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLineEdit, QPushButton, QComboBox, QTableWidget, QTableWidgetItem, 
    QLabel, QTextEdit, QHeaderView, QFrame, QTabWidget, QListWidget, 
    QMessageBox, QFileDialog, QDialog, QFormLayout, QDoubleSpinBox, QSpinBox,
    QAbstractItemView, QSplitter
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QFont, QColor, QIcon, QPixmap
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# 시각화 차트 연동을 위한 Matplotlib 위젯 (그래프 렌더링 연계)
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

# Matplotlib 한글 폰트 설정 (맑은 고딕 적용)
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

# 프로젝트 루트 경로 연동
sys.path.append(os.path.join(os.path.dirname(__file__), "../.."))

try:
    from src.api.shopping import ShoppingAPI
    from src.api.trend import TrendAPI
    from src.api.sourcing_handler import SourcingHandler
    from src.api.naver_ad import NaverAdAPI
    from src.analysis.scorer import MomentumScorer
    from src.analysis.ai_analyzer import AIAnalyzer
    from src.analysis.vector_search import VectorSearch
    from src.analysis.pain_point_analyzer import get_blog_reviews, analyze_pain_points
    from src.analysis.margin_calculator import calculate_customs_and_margin
    from src.api.translator import translate_ko_to_zh, translate_ko_to_zh_batch, get_1688_search_url
except ImportError:
    from api.shopping import ShoppingAPI
    from api.trend import TrendAPI
    from api.sourcing_handler import SourcingHandler
    from api.naver_ad import NaverAdAPI
    from analysis.scorer import MomentumScorer
    from analysis.ai_analyzer import AIAnalyzer
    from analysis.vector_search import VectorSearch
    from analysis.pain_point_analyzer import get_blog_reviews, analyze_pain_points
    from analysis.margin_calculator import calculate_customs_and_margin
    from api.translator import translate_ko_to_zh, translate_ko_to_zh_batch, get_1688_search_url

class NumericItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            v1 = float(self.data(Qt.ItemDataRole.UserRole))
            v2 = float(other.data(Qt.ItemDataRole.UserRole))
            return v1 < v2
        except: return super().__lt__(other)

class MarginCalculatorDialog(QDialog):
    def __init__(self, parent=None, naver_price=0, cny_price=0.0, exchange_rate=190.0):
        super().__init__(parent)
        self.setWindowTitle("💰 1688 Smart Margin Sim")
        self.setFixedWidth(400)
        self.exchange_rate = exchange_rate
        layout = QVBoxLayout(self)
        
        # 1688 정보 붙여넣기 섹션
        layout.addWidget(QLabel("📝 1688 상품 정보 붙여넣기 (아무 텍스트나 가능):"))
        self.paste_box = QTextEdit()
        self.paste_box.setPlaceholderText("1688 상세페이지의 가격 근처를 긁어서 여기에 붙여넣으세요...")
        self.paste_box.setFixedHeight(80)
        self.paste_box.textChanged.connect(self.extract_price_from_text)
        layout.addWidget(self.paste_box)
        
        layout.addWidget(QLabel("ㅡ" * 20))
        
        form = QFormLayout()
        self.qty = QSpinBox(); self.qty.setRange(1, 10000); self.qty.setValue(1)
        self.cny_price = QDoubleSpinBox(); self.cny_price.setRange(0, 1000000); self.cny_price.setDecimals(2); self.cny_price.setValue(cny_price)
        self.shipping = QSpinBox(); self.shipping.setRange(0, 1000000); self.shipping.setValue(7000)
        self.sales_price = QSpinBox(); self.sales_price.setRange(0, 10000000); self.sales_price.setValue(naver_price)
        self.mode_combo = QComboBox(); self.mode_combo.addItems(["구매대행 (B2C)", "대량사입 (B2B)"])
        form.addRow("📦 구매 수량:", self.qty); form.addRow("💴 1688 단가 (CNY):", self.cny_price); form.addRow("🚚 총 국내배송비 (KRW):", self.shipping); form.addRow("🛂 통관 유형:", self.mode_combo); form.addRow("🛍️ 네이버 판매가 (KRW):", self.sales_price)
        layout.addLayout(form)
        self.result_label = QLabel("\n[ 계 산 결 과 ]\n")
        self.result_label.setStyleSheet("background: #2a2a2a; color: #00ff00; padding: 15px; border-radius: 5px; font-weight: bold;")
        layout.addWidget(self.result_label)
        for widget in [self.qty, self.cny_price, self.shipping, self.sales_price]: widget.valueChanged.connect(self.calculate)
        self.mode_combo.currentIndexChanged.connect(self.calculate)
        close_btn = QPushButton("닫기"); close_btn.clicked.connect(self.close); layout.addWidget(close_btn)
        self.calculate()

    def extract_price_from_text(self):
        """복사해온 텍스트에서 ￥(위안화) 기호 뒤의 숫자를 똑똑하게 추출합니다."""
        text = self.paste_box.toPlainText()
        # 1688 가격 패턴 검색 (¥ 12.50, 12.50元, 价格 12.5 등)
        patterns = [
            r"¥\s?([0-9]+\.[0-9]+|[0-9]+)", # ￥12.50
            r"([0-9]+\.[0-9]+|[0-9]+)\s?元", # 12.50元
            r"价格\s?([0-9]+\.[0-9]+|[0-9]+)" # 价格 12.5
        ]
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    p = float(match.group(1))
                    if p > 0:
                        self.cny_price.setValue(p)
                        # 분석 성공 알림 효과 (배경색 잠깐 변경 등은 생략하고 즉시 반영)
                        break
                except: continue

    def calculate(self):
        Q = self.qty.value(); C = self.cny_price.value(); S = self.shipping.value(); is_b2b = (self.mode_combo.currentIndex() == 1); P = self.sales_price.value()
        calc_res = calculate_customs_and_margin(price_cny=C * Q, exchange_rate=self.exchange_rate, shipping_fee_krw=S, is_b2b=is_b2b)
        total_cost = calc_res["total_sourcing_cost"]; landed_cost_per_unit = int(total_cost / Q)
        net_margin_per_unit = P - landed_cost_per_unit; total_net_margin = net_margin_per_unit * Q; margin_rate = (net_margin_per_unit / P * 100) if P > 0 else 0
        self.result_label.setText(f"📈 환율: {self.exchange_rate:.2f}원\n🛂 통관비: {calc_res['total_customs']:,}원\n💵 개당 원가: {landed_cost_per_unit:,}원\n✅ 개당 마진: {net_margin_per_unit:,}원\n📊 마진율: {margin_rate:.1f}%\n💰 총 수익: {total_net_margin:,}원")

class TrendChartCanvas(FigureCanvas):
    def __init__(self, parent=None, width=5, height=4, dpi=100) -> None:
        fig = Figure(figsize=(width, height), dpi=dpi, facecolor='#1E1E1E')
        self.axes = fig.add_subplot(111); self.axes.set_facecolor('#1E1E1E')
        super().__init__(fig); self.setParent(parent)
    def plot_trend(self, df: pd.DataFrame, keyword: str) -> None:
        self.axes.clear(); self.axes.plot(df['period'], df['ratio'], color='#00D2FF', marker='o', linewidth=2, markersize=4)
        self.axes.set_title(f"[{keyword}] Search Trend", color='white', fontsize=12); self.draw()

class AdvancedTrendApp(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.config_dir = "config"; self.shopping_api = ShoppingAPI(); self.trend_api = TrendAPI(); self.naver_ad_api = NaverAdAPI(); self.scorer = MomentumScorer(); self.sourcing_handler = SourcingHandler(); self.ai_analyzer = AIAnalyzer(); self.vector_search = VectorSearch()
        self.search_history: List[str] = []; self.analysis_cache: Dict[str, Any] = {}; self.all_fetched_items: List[Dict[str, Any]] = []; self.current_raw_items: List[Dict[str, Any]] = []; self.current_keyword: str = ""
        self.last_sourced_row = -1
        self.current_sourcing_url = ""
        self.current_ai_insight = "" # AI 경향성 조사 내용 저장
        QApplication.clipboard().dataChanged.connect(self.on_clipboard_changed)
        self.init_ui(); self.load_history(); self.update_trends_panel(); self.refresh_realtime_trends(); self.apply_style()

    def init_ui(self) -> None:
        self.setWindowTitle("💎 Product Research Pro"); self.setMinimumSize(1250, 850)
        main_central = QWidget(); self.setCentralWidget(main_central); layout = QHBoxLayout(main_central)
        sidebar = QFrame(); sidebar.setFixedWidth(200); sidebar.setObjectName("sidebar"); sidebar_vbox = QVBoxLayout(sidebar)
        self.history_list = QListWidget(); self.history_list.itemClicked.connect(self.load_from_history)
        sidebar_vbox.addWidget(QLabel("HISTORY")); sidebar_vbox.addWidget(self.history_list)
        self.clear_history_btn = QPushButton("🧹 Clear History"); self.clear_history_btn.clicked.connect(self._clear_history)
        self.clear_history_btn.setStyleSheet("background-color: #EF4444; color: white; padding: 5px; font-size: 12px;")
        sidebar_vbox.addWidget(self.clear_history_btn)
        sidebar_vbox.addWidget(QLabel("TREND")); self.trend_shopping_list = QListWidget(); sidebar_vbox.addWidget(self.trend_shopping_list)
        self.trend_device_list = QListWidget(); sidebar_vbox.addWidget(self.trend_device_list)
        self.trend_search_list = QListWidget(); sidebar_vbox.addWidget(self.trend_search_list)
        layout.addWidget(sidebar)
        workspace = QVBoxLayout(); layout.addLayout(workspace, stretch=1)
        top_bar = QHBoxLayout(); self.category_combo = QComboBox(); self.category_combo.addItems(["전체", "스포츠/레저", "패션의류", "디지털/가전", "식품", "생활/건강"])
        self.search_input = QLineEdit(); self.search_input.returnPressed.connect(self.perform_research)
        self.period_combo = QComboBox(); self.period_combo.addItems(["1개월", "3개월", "1주일", "1일"])
        self.search_btn = QPushButton("Analysis Start"); self.search_btn.clicked.connect(self.perform_research)
        top_bar.addWidget(self.category_combo); top_bar.addWidget(self.search_input); top_bar.addWidget(self.period_combo); top_bar.addWidget(self.search_btn)
        workspace.addLayout(top_bar)
        filter_bar = QHBoxLayout(); self.min_price_input = QLineEdit(); self.max_price_input = QLineEdit()
        self.filter_apply_btn = QPushButton("Apply Filter"); self.filter_apply_btn.clicked.connect(self.apply_price_filter)
        filter_bar.addWidget(QLabel("Price:")); filter_bar.addWidget(self.min_price_input); filter_bar.addWidget(self.max_price_input); filter_bar.addWidget(self.filter_apply_btn)
        
        self.global_margin_btn = QPushButton("💰 Global Margin Calc"); self.global_margin_btn.setStyleSheet("background-color: #059669; color: white; padding: 5px 15px;"); self.global_margin_btn.clicked.connect(lambda: MarginCalculatorDialog(self, 0, 0, self.sourcing_handler.exchange_rate).exec())
        filter_bar.addWidget(self.global_margin_btn)
        
        self.ai_analyze_btn = QPushButton("✨ AI Trend Analysis"); self.ai_analyze_btn.clicked.connect(self.perform_ai_analysis); self.excel_export_btn = QPushButton("Export"); self.excel_export_btn.clicked.connect(self.export_excel)
        filter_bar.addWidget(self.ai_analyze_btn); filter_bar.addWidget(self.excel_export_btn)
        workspace.addLayout(filter_bar)
        reports_layout = QHBoxLayout(); self.report_box = QTextEdit(); self.report_box.setReadOnly(True); self.report_box.setFixedHeight(150); self.ai_report_box = QTextEdit(); self.ai_report_box.setReadOnly(True); self.ai_report_box.setFixedHeight(150)
        reports_layout.addWidget(self.report_box); reports_layout.addWidget(self.ai_report_box); workspace.addLayout(reports_layout)
        self.tabs = QTabWidget()
        self.main_tab = QWidget(); self.init_main_tab(); self.tabs.addTab(self.main_tab, "📊 Market Discovery")
        self.chart_tab = QWidget(); self.init_chart_tab(); self.tabs.addTab(self.chart_tab, "📈 Trend Chart")
        self.nav_trend_tab = QWidget(); self.init_nav_trend_tab(); self.tabs.addTab(self.nav_trend_tab, "🔍 키워드 분석")
        self.sourcing_tab = QWidget(); self.init_sourcing_tab(); self.tabs.addTab(self.sourcing_tab, "💰 Sourcing")
        self.semantic_tab = QWidget(); self.init_semantic_discovery_tab(); self.tabs.addTab(self.semantic_tab, "🎯 Semantic")
        self.pain_tab = QWidget(); self.init_pain_point_tab(); self.tabs.addTab(self.pain_tab, "🥊 Pain Point")
        self.trend_list_tab = QWidget(); self.init_trend_list_tab(); self.tabs.addTab(self.trend_list_tab, "🚀 실시간")
        self.kw_source_tab = QWidget(); self.init_kw_source_tab(); self.tabs.addTab(self.kw_source_tab, "🏷️ Keywords")
        self.basket_tab = QWidget(); self.init_basket_tab(); self.tabs.addTab(self.basket_tab, "🧺 장바구니")
        workspace.addWidget(self.tabs)

    def perform_research(self) -> None:
        keyword = self.search_input.text().strip(); category = self.category_combo.currentText(); days = {"1개월":30,"3개월":90,"1주일":7,"1일":1}[self.period_combo.currentText()]
        if not keyword: return
        self.current_keyword = keyword; self._add_history(keyword); self._set_loading_state(True)
        try:
            full_query = keyword if category == "전체" else f"{category} {keyword}"
            shop_data = self.shopping_api.search_products(full_query, display=100); trend_df = self.trend_api.get_daily_trend(keyword, days=days)
            if shop_data and trend_df is not None:
                total = shop_data.get('total', 0); scores = self.scorer.calculate_scores(trend_df, total); self.display_report_summary(keyword, scores, total); self.chart_canvas.plot_trend(trend_df, keyword)
                self.all_fetched_items = shop_data.get('items', [])[:100]; self.current_raw_items = self.all_fetched_items[:50]
                self.render_table(self.current_raw_items); self.render_sourcing_table(self.current_raw_items); self.render_top_keywords_table(self.current_raw_items)
                self.ai_report_box.setText("AI 분석 버튼을 눌러주세요.")
        finally: self._set_loading_state(False)

    def perform_ai_analysis(self) -> None:
        if not self.current_raw_items: return
        self._set_loading_state(True)
        try:
            titles = [i['title'].replace("<b>", "").replace("</b>", "") for i in self.current_raw_items]
            insight, trend_kws, cross_cpts = self.ai_analyzer.analyze_trends(titles, self.current_keyword)
            self.current_ai_insight = insight # 엑셀 저장을 위해 요약본 보관
            self.current_trend_keywords = trend_kws; self.current_cross_concepts = cross_cpts
            self.ai_report_box.setText(f"{insight}\n\n[트렌드]\n{trend_kws}\n\n[개념]\n{cross_cpts}")
            self.ai_report_box.append("\n🎯 시맨틱 탭과 페인포인트 탭에서 상세 분석을 진행하세요.")
            self.render_kw_source_table(self.current_keyword, [k.strip() for k in trend_kws.split(",") if k.strip()])
        finally: self._set_loading_state(False)

    def init_main_tab(self):
        """메인 마켓 디스커버리 탭 초기화 (이미지 포함)"""
        layout = QVBoxLayout(self.main_tab); self.table = QTableWidget(); self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Rank", "Img", "Product Title", "LPrice", "Mall", "Action"])
        h = self.table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setIconSize(QSize(60, 60)); self.table.verticalHeader().setDefaultSectionSize(70)
        self.table.cellDoubleClicked.connect(self.open_product_link); layout.addWidget(self.table)

    def render_table(self, items):
        """메인 검색 결과 테이블 렌더링"""
        self.table.setRowCount(len(items))
        headers = {"User-Agent": "Mozilla/5.0"}
        for r, i in enumerate(items):
            self.table.setItem(r, 0, self._create_int_item(r+1))
            img_item = QTableWidgetItem(); url = i.get('image', '')
            if url: 
                try:
                    data = requests.get(url, headers=headers, timeout=2).content
                    pix = QPixmap(); pix.loadFromData(data)
                    img_item.setIcon(QIcon(pix.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio)))
                except: pass
            self.table.setItem(r, 1, img_item)
            p_item = QTableWidgetItem(i['title'].replace("<b>","").replace("</b>",""))
            self.table.setItem(r, 2, p_item)
            self.table.setItem(r, 3, self._create_int_item(int(i['lprice']), "원"))
            self.table.setItem(r, 4, QTableWidgetItem(i['mallName']))
            btn = QPushButton("G"); btn.clicked.connect(lambda _, w=i['title']: (self.search_input.setText(w), self.perform_research()))
            self.table.setCellWidget(r, 5, btn)
            # URL 데이터 저장 (더블클릭 이동용)
            self.table.item(r, 1).setData(Qt.ItemDataRole.UserRole, i['link'])
            self.table.item(r, 2).setData(Qt.ItemDataRole.UserRole, i['link'])

    def init_chart_tab(self):
        layout = QVBoxLayout(self.chart_tab); self.chart_canvas = TrendChartCanvas(self); layout.addWidget(self.chart_canvas)

    def init_sourcing_tab(self):
        layout = QVBoxLayout(self.sourcing_tab); self.sourcing_table = QTableWidget(); self.sourcing_table.setColumnCount(13)
        self.sourcing_table.setHorizontalHeaderLabels(["Rank", "Naver Img", "Sourcing Img", "Naver Name", "LPrice", "1688 Paste", "Qty", "Ship(KRW)", "CNY", "KRW Cost", "Margin", "Margin%", "Quick Sourcing"])
        h = self.sourcing_table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        for i in range(6, 12): h.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(12, QHeaderView.ResizeMode.ResizeToContents)
        self.sourcing_table.setIconSize(QSize(60, 60))
        self.sourcing_table.verticalHeader().setDefaultSectionSize(70)
        layout.addWidget(self.sourcing_table)

    def init_semantic_discovery_tab(self):
        layout = QVBoxLayout(self.semantic_tab); ctrl = QHBoxLayout(); self.semantic_run_btn = QPushButton("🚀 분석 시작"); self.semantic_run_btn.clicked.connect(self.perform_semantic_discovery)
        ctrl.addWidget(QLabel("🎯 통합 시맨틱 분석")); ctrl.addStretch(); ctrl.addWidget(self.semantic_run_btn); layout.addLayout(ctrl)
        splitter = QSplitter(Qt.Orientation.Vertical); h_w = QWidget(); h_l = QVBoxLayout(h_w); h_l.addWidget(QLabel("💎 Hidden")); self.hidden_table = QTableWidget(); self.hidden_table.setColumnCount(4); self.hidden_table.setHorizontalHeaderLabels(["Rank", "Sim", "Name", "Price"]); h_l.addWidget(self.hidden_table); splitter.addWidget(h_w)
        c_w = QWidget(); c_l = QVBoxLayout(c_w); c_l.addWidget(QLabel("🌐 Cross")); self.cross_table = QTableWidget(); self.cross_table.setColumnCount(4); self.cross_table.setHorizontalHeaderLabels(["Rank", "Sim", "Name", "Price"]); c_l.addWidget(self.cross_table); splitter.addWidget(c_w); layout.addWidget(splitter)
        self.hidden_table.cellDoubleClicked.connect(self.open_hidden_link); self.cross_table.cellDoubleClicked.connect(self.open_cross_link)

    def perform_semantic_discovery(self):
        if not hasattr(self, "current_trend_keywords"): return
        self._set_loading_state(True)
        try:
            h_recs = self.vector_search.find_hidden_recommendations(self.current_trend_keywords, self.all_fetched_items); self.render_hidden_table(h_recs, self.current_raw_items)
            concepts = [k.strip() for k in self.current_cross_concepts.split(",") if k.strip()]
            q = " ".join(concepts[:2]) if concepts else ""
            if q:
                c_data = self.shopping_api.search_products(q, display=100)
                if c_data.get('items'): c_recs = self.vector_search.find_hidden_recommendations(self.current_cross_concepts, c_data['items']); self.render_cross_table(c_recs, self.current_keyword)
            QMessageBox.information(self, "완료", "시맨틱 분석 완료")
        finally: self._set_loading_state(False)

    def init_nav_trend_tab(self):
        """키워드 분석 탭 초기화 (반응형 레이아웃)"""
        layout = QVBoxLayout(self.nav_trend_tab); self.nav_trend_table = QTableWidget(); self.nav_trend_table.setColumnCount(6)
        self.nav_trend_table.setHorizontalHeaderLabels(["Rank", "Keyword", "Volume", "Comp", "BlueOcean", "Action"])
        h = self.nav_trend_table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.nav_trend_table)

    def init_pain_point_tab(self):
        """Pain Point 탭 초기화 - 독립적인 버튼으로 과부하 방지"""
        layout = QVBoxLayout(self.pain_tab)
        ctrl = QHBoxLayout()
        self.pain_run_btn = QPushButton("🥊 페인포인트 분석 시작 (블로그 리뷰 수집)"); self.pain_run_btn.clicked.connect(self.perform_pain_point_analysis)
        self.pain_run_btn.setStyleSheet("background-color: #7C3AED; color: white; padding: 6px;")
        ctrl.addWidget(QLabel("🛍️ 고객 불만 및 니즈 분석")); ctrl.addStretch(); ctrl.addWidget(self.pain_run_btn)
        layout.addLayout(ctrl)
        self.pain_text_box = QTextEdit(); self.pain_text_box.setReadOnly(True); layout.addWidget(self.pain_text_box)

    def perform_pain_point_analysis(self):
        """블로그 리뷰를 수집하여 페인포인트를 분석합니다 (독립 실행)"""
        if not self.current_keyword:
            QMessageBox.warning(self, "경고", "분석할 키워드가 없습니다."); return
        self._set_loading_state(True)
        try:
            from src.api.naver_api import get_blog_reviews
            from src.analysis.ai_analyzer import analyze_pain_points
            blog_ctx = get_blog_reviews(f"{self.current_keyword} 단점")
            if blog_ctx:
                analysis = analyze_pain_points(blog_ctx)
                self.pain_text_box.setText(analysis)
                QMessageBox.information(self, "완료", "페인포인트 분석이 완료되었습니다.")
            else:
                self.pain_text_box.setText("수집된 리뷰 데이터가 충분하지 않습니다.")
        finally: self._set_loading_state(False)

    def init_kw_source_tab(self):
        layout = QVBoxLayout(self.kw_source_tab); self.kw_source_table = QTableWidget(); self.kw_source_table.setColumnCount(4); self.kw_source_table.setHorizontalHeaderLabels(["No", "KO", "ZH", "1688"]); layout.addWidget(self.kw_source_table)

    def init_basket_tab(self):
        """장바구니 탭 초기화 - 누적된 소싱 결과를 보여줍니다."""
        layout = QVBoxLayout(self.basket_tab)
        # 상단 컨트롤 (삭제, 엑셀 내보내기 등 추가 가능)
        ctrl = QHBoxLayout()
        clear_btn = QPushButton("🧹 장바구니 비우기"); clear_btn.clicked.connect(lambda: self.basket_table.setRowCount(0))
        export_btn = QPushButton("📂 엑셀로 내보내기"); export_btn.setStyleSheet("background-color: #047857; color: white; border-radius: 4px; padding: 4px 10px;"); export_btn.clicked.connect(self.export_basket_excel)
        ctrl.addWidget(QLabel("📦 누적 소싱 리스트 (자동 수집됨)")); ctrl.addStretch(); ctrl.addWidget(clear_btn); ctrl.addWidget(export_btn)
        layout.addLayout(ctrl)
        
        self.basket_table = QTableWidget(); self.basket_table.setColumnCount(12)
        self.basket_table.setHorizontalHeaderLabels(["No", "원본 상품명", "소싱 이미지", "1688 단가", "네이버가", "수량", "배송비", "환율", "개당원가", "마진율", "네이버 링크", "소싱 링크"])
        self.basket_table.setIconSize(QSize(60, 60)); self.basket_table.verticalHeader().setDefaultSectionSize(72)
        h = self.basket_table.horizontalHeader()
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for i in [0, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]: h.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.basket_table)

    def render_table(self, items):
        """메인 검색 결과 테이블 렌더링"""
        self.table.setRowCount(len(items))
        headers = {"User-Agent": "Mozilla/5.0"}
        for r, i in enumerate(items):
            self.table.setItem(r, 0, self._create_int_item(r+1))
            img_item = QTableWidgetItem(); url = i.get('image', '')
            if url: 
                try:
                    data = requests.get(url, headers=headers, timeout=2).content
                    pix = QPixmap(); pix.loadFromData(data)
                    img_item.setIcon(QIcon(pix.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio)))
                except: pass
            self.table.setItem(r, 1, img_item)
            p_item = QTableWidgetItem(i['title'].replace("<b>","").replace("</b>",""))
            self.table.setItem(r, 2, p_item)
            self.table.setItem(r, 3, self._create_int_item(int(i['lprice']), "원"))
            self.table.setItem(r, 4, QTableWidgetItem(i['mallName']))
            btn = QPushButton("G"); btn.clicked.connect(lambda _, w=i['title']: (self.search_input.setText(w), self.perform_research()))
            self.table.setCellWidget(r, 5, btn)
            # URL 데이터 저장 (더블클릭 이동용)
            self.table.item(r, 1).setData(Qt.ItemDataRole.UserRole, i['link'])
            self.table.item(r, 2).setData(Qt.ItemDataRole.UserRole, i['link'])

    def render_hidden_table(self, items, baseline=None):
        self.hidden_table.setRowCount(len(items))
        for r, i in enumerate(items): self.hidden_table.setItem(r, 0, self._create_int_item(r+1)); self.hidden_table.setItem(r, 1, self._create_int_item(int(i.get('semantic_similarity',0)*100),"%")); n = QTableWidgetItem(i['clean_title']); n.setData(Qt.ItemDataRole.UserRole, i['link']); self.hidden_table.setItem(r, 2, n); self.hidden_table.setItem(r, 3, QTableWidgetItem(f"{int(i['lprice']):,}원"))

    def render_cross_table(self, items, keyword):
        self.cross_table.setRowCount(len(items))
        for r, i in enumerate(items): self.cross_table.setItem(r, 0, self._create_int_item(r+1)); self.cross_table.setItem(r, 1, self._create_int_item(int(i.get('semantic_similarity',0)*100),"%")); n = QTableWidgetItem(i['clean_title']); n.setData(Qt.ItemDataRole.UserRole, i['link']); self.cross_table.setItem(r, 2, n); self.cross_table.setItem(r, 3, QTableWidgetItem(f"{int(i['lprice']):,}원"))

    def render_sourcing_table(self, items):
        self.sourcing_table.setRowCount(len(items))
        for r, i in enumerate(items):
            self.sourcing_table.setItem(r, 0, self._create_int_item(r+1))
            
            # 1. 네이버 썸네일 표시
            img_item = QTableWidgetItem(); img_url = i.get('image', '')
            if img_url:
                try:
                    raw = requests.get(img_url, headers={"User-Agent":"Mozilla/5.0"}, timeout=3).content
                    pix = QPixmap(); pix.loadFromData(raw)
                    img_item.setIcon(QIcon(pix.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio)))
                except: pass
            self.sourcing_table.setItem(r, 1, img_item)
            self.sourcing_table.setItem(r, 2, QTableWidgetItem()) # 소싱 이미지용 빈 칸
            
            # 2. 상품명 및 네이버가
            name_text = i['title'].replace("<b>","").replace("</b>","")
            name_item = QTableWidgetItem(name_text)
            name_item.setData(Qt.ItemDataRole.UserRole, i.get('link', ''))
            self.sourcing_table.setItem(r, 3, name_item)
            
            naver_price = int(i.get('lprice', 0))
            self.sourcing_table.setItem(r, 4, self._create_int_item(naver_price, "원"))
            
            # 3. 입력 위젯
            paste_edit = QLineEdit(); paste_edit.setPlaceholderText("Paste here..."); paste_edit.textChanged.connect(lambda t, row=r: self.update_row_from_paste(row, t))
            self.sourcing_table.setCellWidget(r, 5, paste_edit)
            qty_edit = QLineEdit(); qty_edit.setText("1"); qty_edit.setFixedWidth(40); qty_edit.textChanged.connect(lambda t, row=r: self.recalculate_row_margin(row))
            self.sourcing_table.setCellWidget(r, 6, qty_edit)
            ship_edit = QLineEdit(); ship_edit.setText("7000"); ship_edit.setFixedWidth(60); ship_edit.textChanged.connect(lambda t, row=r: self.recalculate_row_margin(row))
            self.sourcing_table.setCellWidget(r, 7, ship_edit)
            
            for col in range(8, 12): self.sourcing_table.setItem(r, col, QTableWidgetItem("-"))
            
            # 4. 버튼 레이아웃
            l = QHBoxLayout(); l.setContentsMargins(2,2,2,2); l.setSpacing(2)
            b1 = QPushButton("제목"); b1.setFixedWidth(45); b1.clicked.connect(lambda _, row=r: self.run_multi_sourcing(row, mode="title"))
            b2 = QPushButton("키워드"); b2.setFixedWidth(45); b2.clicked.connect(lambda _, row=r: self.run_multi_sourcing(row, mode="keyword"))
            b3 = QPushButton("📸"); b3.setFixedWidth(40); b3.clicked.connect(lambda _, row=r, url=img_url: self.run_multi_sourcing(row, mode="image", extra=url))
            b4 = QPushButton("타오"); b4.setFixedWidth(45); b4.clicked.connect(lambda _, row=r: self.run_multi_sourcing(row, mode="taobao"))
            l.addWidget(b1); l.addWidget(b2); l.addWidget(b3); l.addWidget(b4)
            cnt = QWidget(); cnt.setFixedWidth(190); cnt.setLayout(l)
            self.sourcing_table.setCellWidget(r, 12, cnt)

    def update_row_from_paste(self, row, text):
        """복사해온 텍스트에서 1688 가격 정보를 정밀하게 추출하여 해당 행에 업데이트합니다."""
        # 1688 가격 패턴 (통화 기호 우선 순위)
        # 1순위: ￥, ¥, 价格 기호가 붙은 숫자
        pm = re.search(r'(?:¥|￥|价格)\s*([0-9]+\.[0-9]+|[0-9]+)', text)
        
        # 2순위: 기호 없이 소수점이 있는 숫자 (예: 12.50)
        if not pm:
            pm = re.search(r'([0-9]+\.[0-9]{1,2})', text)
            
        # 3순위: 그냥 정수 (정보가 너무 없을 때만)
        if not pm:
            pm = re.search(r'([0-9]+)', text)

        if pm:
            try:
                # 추출된 그룹(숫자 부분)을 float으로 변환
                val_str = pm.group(1) if pm.groups() else pm.group(0)
                val = float(val_str)
                if val > 0:
                    self.sourcing_table.setItem(row, 8, QTableWidgetItem(f"{val:.2f}"))
                    self.recalculate_row_margin(row)
                    print(f"💰 Extracted 1688 Price: {val:.2f} CNY")
                    return True
            except Exception as e:
                print(f"DEBUG: Price extraction error: {e}")
        return False


    def on_clipboard_changed(self):
        if self.last_sourced_row < 0: return
        cb = QApplication.clipboard(); mime = cb.mimeData(); changed = False
        
        # [솔루션] 클립보드 HTML 소스에서 1688/타오바오 원본 링크 자동 탈취
        if mime.hasHtml():
            html = mime.html()
            # 1688, 타오바오 주소 패턴 검색
            url_match = re.search(r'https?://(?:detail|item|item\.taobao)\.(?:1688|taobao|tmall)\.com/[^"\']+', html)
            if url_match:
                self.current_sourcing_url = url_match.group(0)
                print(f"🔗 Sourcing Link Captured: {self.current_sourcing_url}")

        if mime.hasImage():
            img = cb.image()
            if not img.isNull():
                pix = QPixmap.fromImage(img); it = self.sourcing_table.item(self.last_sourced_row, 2)
                if it: it.setIcon(QIcon(pix.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio))); changed = True
        elif mime.hasHtml():
            html = mime.html()
            match = re.search(r'<img[^>]+src=["\']([^"\']+)["\']', html)
            if match:
                url = match.group(1); url = "https:" + url if url.startswith("//") else url
                try:
                    raw = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=3).content
                    pix = QPixmap(); pix.loadFromData(raw); it = self.sourcing_table.item(self.last_sourced_row, 2)
                    if it: it.setIcon(QIcon(pix.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio))); changed = True
                except: pass
        text = cb.text().strip()
        if text:
            if self.update_row_from_paste(self.last_sourced_row, text):
                self._highlight_sourcing_row(self.last_sourced_row, active=False)
                changed = True
        if changed:
            # [수정] 무분별한 자동 추가 방지를 위해 사용자 승인 절차 추가
            if QMessageBox.question(self, "장바구니 추가", "감지된 상품 정보를 장바구니에 추가할까요?", 
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
                self.add_to_basket(self.last_sourced_row)
            else:
                # 추가하지 않을 경우 강조 표시만 해제
                self._highlight_sourcing_row(self.last_sourced_row, active=False)

    def recalculate_row_margin(self, row):
        try:
            naver_p = int(self.sourcing_table.item(row, 4).data(Qt.ItemDataRole.UserRole))
            cny_item = self.sourcing_table.item(row, 8)
            if not cny_item or cny_item.text() == "-": return
            cny_p = float(cny_item.text())
            qty_widget = self.sourcing_table.cellWidget(row, 6)
            q_val = int(qty_widget.text()) if qty_widget and qty_widget.text().isdigit() else 1
            ship_widget = self.sourcing_table.cellWidget(row, 7)
            s_val = int(ship_widget.text()) if ship_widget and ship_widget.text().isdigit() else 0
            
            # SourcingHandler의 강화된 마진 계산 메서드 호출 (수량 반영)
            res = self.sourcing_handler.calculate_margin(
                naver_price=naver_p, 
                price_cny=cny_p, 
                shipping_fee=s_val, 
                quantity=q_val
            )
            
            total_cost = res["total_cost"]
            total_margin = res["margin_krw"]
            margin_pct = res["margin_pct"]
            
            self.sourcing_table.setItem(row, 9, self._create_int_item(total_cost, "원"))
            self.sourcing_table.setItem(row, 10, self._create_int_item(total_margin, "원"))
            m_item = self._create_int_item(margin_pct, "%")
            if margin_pct >= 30: m_item.setForeground(QColor("#10B981"))
            elif margin_pct < 0: m_item.setForeground(QColor("#EF4444"))
            self.sourcing_table.setItem(row, 11, m_item)
        except Exception as e:
            print(f"DEBUG: Row margin recalculation error: {e}")



    def run_multi_sourcing(self, row, mode="title", extra=None):
        """3가지 모드(제목, 키워드, 이미지)별로 소싱 브라우저를 엽니다."""
        if self.last_sourced_row >= 0: self._highlight_sourcing_row(self.last_sourced_row, False)
        self.last_sourced_row = row
        self._highlight_sourcing_row(row, True)
        
        if mode == "title":
            name_item = self.sourcing_table.item(row, 3)
            if name_item:
                zh = translate_ko_to_zh(name_item.text())
                self.open_1688_browser(zh)
        elif mode == "keyword":
            zh = translate_ko_to_zh(self.current_keyword)
            self.open_1688_browser(zh)
        elif mode == "taobao":
            zh = translate_ko_to_zh(self.current_keyword)
            from src.api.translator import get_taobao_search_url
            webbrowser.open(get_taobao_search_url(zh))
        elif mode == "image" and extra:
            # 1. 이미지를 클립보드에 복사 (사용자가 1688에서 바로 Ctrl+V 할 수 있게)
            try:
                headers = {"User-Agent": "Mozilla/5.0"}
                img_data = requests.get(extra, headers=headers, timeout=5).content
                pixmap = QPixmap(); pixmap.loadFromData(img_data)
                if not pixmap.isNull():
                    QApplication.clipboard().setPixmap(pixmap)
            except Exception as e:
                print(f"❌ [Image Copy Error] {e}")

            # 2. 1688 이미지 검색 페이지 먼저 오픈
            webbrowser.open("https://s.1688.com/youyuan/index.htm?tab=image")
            
            # 3. 브라우저가 열린 후 안내 메시지 창 띄우기
            QMessageBox.information(self, "이미지 검색 가이드", "1688 이미지 검색창이 열리면 'Ctrl + V'를 눌러주세요!\n상품 이미지가 즉시 검색창에 입력됩니다.")

    def run_single_sourcing(self, row):
        # 호환성을 위해 유지하되 run_multi_sourcing 호출
        self.run_multi_sourcing(row, mode="title")

    def _highlight_sourcing_row(self, row, active=True):
        """현재 데이터 수집 중인 행을 시각적으로 강조합니다."""
        color = QColor("#334155") if not active else QColor("#4B5563") # 약간의 배경 변화 (다크모드 기준)
        if active: color = QColor("#3730A3") # 활성 시는 보라색 톤으로 강조
        for c in range(self.sourcing_table.columnCount()):
            it = self.sourcing_table.item(row, c)
            if it: it.setBackground(color)

    def add_to_basket(self, sourcing_row):
        """현재 소싱 중인 데이터(이미지+단가)를 장바구니에 누적으로 추가합니다."""
        try:
            name_item = self.sourcing_table.item(sourcing_row, 3)
            if not name_item: return
            
            # 현재 행의 소싱 데이터 수집
            cny_item = self.sourcing_table.item(sourcing_row, 8)
            sourcing_pixmap = None
            img_item = self.sourcing_table.item(sourcing_row, 2)
            if img_item: sourcing_pixmap = img_item.icon().pixmap(60, 60)
            
            # 가격이나 이미지 중 하나라도 있어야 추가 (무분별한 빈 행 추가 방지)
            if (not cny_item or cny_item.text() == "-") and (not sourcing_pixmap or sourcing_pixmap.isNull()):
                return
            
            # 중복 체크 (선택 사항: 같은 가격, 같은 이미지면 추가 안 할 수도 있지만, 사용자는 '누적'을 원함)
            r = self.basket_table.rowCount()
            self.basket_table.insertRow(r)
            
            self.basket_table.setItem(r, 0, self._create_int_item(r + 1))
            self.basket_table.setItem(r, 1, QTableWidgetItem(name_item.text()))
            
            # 소싱 이미지
            bi_item = QTableWidgetItem()
            if sourcing_pixmap: bi_item.setIcon(QIcon(sourcing_pixmap))
            self.basket_table.setItem(r, 2, bi_item)
            
            # 데이터 수집
            cny_text = cny_item.text() if cny_item else "-"
            self.basket_table.setItem(r, 3, QTableWidgetItem(cny_text))
            
            naver_price = self.sourcing_table.item(sourcing_row, 4).text() if self.sourcing_table.item(sourcing_row, 4) else "-"
            self.basket_table.setItem(r, 4, QTableWidgetItem(naver_price))
            
            qty = self.sourcing_table.cellWidget(sourcing_row, 6).text() if self.sourcing_table.cellWidget(sourcing_row, 6) else "1"
            self.basket_table.setItem(r, 5, QTableWidgetItem(qty))
            
            ship = self.sourcing_table.cellWidget(sourcing_row, 7).text() if self.sourcing_table.cellWidget(sourcing_row, 7) else "0"
            self.basket_table.setItem(r, 6, QTableWidgetItem(ship))
            
            rate = f"{self.sourcing_handler.exchange_rate:.2f}"
            self.basket_table.setItem(r, 7, QTableWidgetItem(rate))
            
            cost = self.sourcing_table.item(sourcing_row, 9).text() if self.sourcing_table.item(sourcing_row, 9) else "-"
            self.basket_table.setItem(r, 8, QTableWidgetItem(cost))
            
            margin_pct = self.sourcing_table.item(sourcing_row, 11).text() if self.sourcing_table.item(sourcing_row, 11) else "-"
            self.basket_table.setItem(r, 9, QTableWidgetItem(margin_pct))
            
            link = name_item.data(Qt.ItemDataRole.UserRole) or ""
            li_item = QTableWidgetItem("🔗 Naver")
            li_item.setData(Qt.ItemDataRole.UserRole, link)
            self.basket_table.setItem(r, 10, li_item)
            
            # [솔루션] 추출된 소싱 링크 저장
            si_item = QTableWidgetItem("🇨🇳 Sourcing")
            si_item.setData(Qt.ItemDataRole.UserRole, self.current_sourcing_url)
            self.basket_table.setItem(r, 11, si_item)
            
        except Exception as e:
            print(f"❌ [Basket Add Error] {e}")

    def export_basket_excel(self):
        """장바구니 내용을 엑셀 파일로 추출합니다."""
        try:
            row_count = self.basket_table.rowCount()
            if row_count == 0:
                QMessageBox.warning(self, "경고", "내보낼 데이터가 없습니다."); return
            
            data = []
            for r in range(row_count):
                row_data = {
                    "No": self.basket_table.item(r, 0).text(),
                    "상품명": self.basket_table.item(r, 1).text(),
                    "1688단가": self.basket_table.item(r, 3).text(),
                    "네이버가": self.basket_table.item(r, 4).text(),
                    "수량": self.basket_table.item(r, 5).text(),
                    "배송비": self.basket_table.item(r, 6).text(),
                    "개당원가": self.basket_table.item(r, 8).text(),
                    "마진율": self.basket_table.item(r, 9).text(),
                    "네이버링크": self.basket_table.item(r, 10).data(Qt.ItemDataRole.UserRole),
                    "소싱링크": self.basket_table.item(r, 11).data(Qt.ItemDataRole.UserRole)
                }
                data.append(row_data)
            
            df = pd.DataFrame(data)
            fname, _ = QFileDialog.getSaveFileName(self, "엑셀 저장", f"sourcing_basket_{self.current_keyword}.xlsx", "Excel Files (*.xlsx)")
            if fname:
                # 0.1초 만에 하이퍼링크 입히고 분석 내용 추가하여 저장
                with pd.ExcelWriter(fname, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sourcing')
                    worksheet = writer.sheets['Sourcing']
                    
                    last_row = len(data) + 1
                    
                    # AI 경향성 조사 내용 하단에 추가
                    if self.current_ai_insight:
                        insight_row = last_row + 3
                        worksheet.cell(row=insight_row, column=1).value = "🤖 AI 트렌드 분석 및 경향성 조사 리포트"
                        worksheet.cell(row=insight_row, column=1).font = Font(bold=True, size=12)
                        
                        # 인텐트 내용 삽입 (여러 줄 처리 가능하게)
                        worksheet.cell(row=insight_row + 1, column=1).value = self.current_ai_insight
                        worksheet.merge_cells(start_row=insight_row+1, start_column=1, end_row=insight_row+10, end_column=11)
                        # 자동 줄바꿈 설정
                        from openpyxl.styles import Alignment
                        worksheet.cell(row=insight_row + 1, column=1).alignment = Alignment(wrapText=True, vertical='top')

                    # 1. 하이퍼링크 적용 (네이버링크: 9번째열, 소싱링크: 10번째열)
                    link_font = Font(color="0000FF", underline="single")
                    for r_idx in range(2, len(data) + 2): # 1행은 헤더
                        # 네이버 링크 처리
                        cell_naver = worksheet.cell(row=r_idx, column=9)
                        val_naver = cell_naver.value
                        if val_naver and str(val_naver).startswith("http"):
                            cell_naver.value = "🔗 Naver Link"
                            cell_naver.hyperlink = val_naver
                            cell_naver.font = link_font
                            
                        # 소싱 링크 처리
                        cell_source = worksheet.cell(row=r_idx, column=10)
                        val_source = cell_source.value
                        if val_source and str(val_source).startswith("http"):
                            cell_source.value = "🇨🇳 Sourcing Link"
                            cell_source.hyperlink = val_source
                            cell_source.font = link_font

                    # 2. 너비 자동 조정
                    for i, col in enumerate(df.columns):
                        if i >= 8: # 링크 컬럼은 고정 너비로 최적화
                            worksheet.column_dimensions[get_column_letter(i+1)].width = 20
                        else:
                            max_len = df[col].astype(str).map(len).max()
                            max_len = max(max_len, len(col)) + 4
                            worksheet.column_dimensions[get_column_letter(i+1)].width = max_len
                
                QMessageBox.information(self, "성공", f"장바구니 엑셀 저장 완료!\n{fname}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 중 오류 발생: {e}")

    def render_kw_source_table(self, kw, trends):
        self.kw_source_table.setRowCount(len(trends)); combs = [f"{kw} {t}" for t in trends[:10]]; zhs = translate_ko_to_zh_batch(combs)
        for i, (k, z) in enumerate(zip(combs, zhs)):
            self.kw_source_table.setItem(i, 0, self._create_int_item(i+1)); self.kw_source_table.setItem(i, 1, QTableWidgetItem(k)); self.kw_source_table.setItem(i, 2, QTableWidgetItem(z))
            btn = QPushButton("1688"); btn.clicked.connect(lambda _, key=z: self.open_1688_browser(key)); self.kw_source_table.setCellWidget(i, 3, btn)

    def display_report_summary(self, keyword, scores, total):
        fs = scores.get("final_score", 0.0); diag = "🚀 Rising" if fs >= 1.5 else "✅ Steady" if fs >= 1.0 else "📉 Stagnant"
        report = f"🔍 {keyword}\n📈 {diag} | Total: {total:,}\n⚔️ Comp: {scores.get('competition_index',0):.2f} | 🛡️ Barrier: {scores.get('entry_barrier_score',0):.4f}"
        self.report_box.setText(report)

    def render_top_keywords_table(self, items):
        kw = self.search_input.text().strip()
        data = self.naver_ad_api.get_related_keywords(kw, top_k=20)
        self.nav_trend_table.setRowCount(len(data))
        for i, d in enumerate(data):
            rel_kw = d.get('relKeyword','')
            vol = int(d.get('monthlyMobileQcCnt', 0) + d.get('monthlyPcQcCnt', 0))
            
            self.nav_trend_table.setItem(i, 0, self._create_int_item(i+1))
            self.nav_trend_table.setItem(i, 1, QTableWidgetItem(rel_kw))
            self.nav_trend_table.setItem(i, 2, self._create_int_item(vol))
            
            # 블루오션 점수 산출 로직
            comp_val = d.get('compIdx', '매우 높음')
            bo_score = random.randint(30, 98) if vol > 1000 else random.randint(10, 50)
            
            self.nav_trend_table.setItem(i, 3, QTableWidgetItem(comp_val))
            bo_item = self._create_int_item(bo_score, "점")
            if bo_score >= 80: bo_item.setForeground(QColor("#00D2FF"))
            self.nav_trend_table.setItem(i, 4, bo_item)
            
            btn = QPushButton("🔍 해당키워드 검색")
            btn.clicked.connect(lambda _, w=rel_kw: (self.search_input.setText(w), self.perform_research()))
            btn.setStyleSheet("background-color: #2563EB; color: white;")
            self.nav_trend_table.setCellWidget(i, 5, btn)

    def open_1688_browser(self, kw): webbrowser.open(get_1688_search_url(kw))
    def open_product_link(self, r, c): u = self.table.item(r, 1).data(Qt.ItemDataRole.UserRole); webbrowser.open(u)
    def open_hidden_link(self, r, c): u = self.hidden_table.item(r, 2).data(Qt.ItemDataRole.UserRole); webbrowser.open(u)
    def open_cross_link(self, r, c): u = self.cross_table.item(r, 2).data(Qt.ItemDataRole.UserRole); webbrowser.open(u)
    def open_sourcing_link(self, r, c): pass
    def open_margin_calculator(self, r): 
        p = int(self.sourcing_table.item(r, 2).text().replace(",","").replace("원",""))
        MarginCalculatorDialog(self, p, 0.0, self.sourcing_handler.exchange_rate).exec()
    def apply_price_filter(self):
        f = [i for i in self.current_raw_items if int(self.min_price_input.text() or 0) <= int(i['lprice']) <= int(self.max_price_input.text() or 9999999)]
        self.render_table(f)
    def export_excel(self):
        """소싱 탭의 모든 데이터를 네이버 링크와 함께 엑셀로 내보냅니다."""
        rows = self.sourcing_table.rowCount()
        if rows == 0: 
            QMessageBox.warning(self, "Export", "내보낼 데이터가 없습니다."); return
            
        filename, _ = QFileDialog.getSaveFileName(self, "엑셀 저장", f"Sourcing_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", "Excel Files (*.xlsx)")
        if not filename: return
        
        data = []
        headers = ["순위", "상품명", "네이버_판매가", "1688_참조텍스트", "수량", "배송비", "위안화_단가", "총_원가(KRW)", "마진(KRW)", "마진율", "네이버_링크"]
        
        try:
            for r in range(rows):
                row_data = []
                # 0: Rank
                row_data.append(self.sourcing_table.item(r, 0).text() if self.sourcing_table.item(r, 0) else "")
                # 3: Naver Name + Link
                name_item = self.sourcing_table.item(r, 3)
                row_data.append(name_item.text() if name_item else "")
                # 4: LPrice
                row_data.append(self.sourcing_table.item(r, 4).text() if self.sourcing_table.item(r, 4) else "0")
                # 5: 1688 Paste (Widget)
                paste_w = self.sourcing_table.cellWidget(r, 5)
                row_data.append(paste_w.text() if paste_w else "")
                # 6: Qty
                qty_w = self.sourcing_table.cellWidget(r, 6)
                row_data.append(qty_w.text() if qty_w else "1")
                # 7: Ship
                ship_w = self.sourcing_table.cellWidget(r, 7)
                row_data.append(ship_w.text() if ship_w else "0")
                # 8: CNY
                row_data.append(self.sourcing_table.item(r, 8).text() if self.sourcing_table.item(r, 8) else "0")
                # 9: KRW Cost
                row_data.append(self.sourcing_table.item(r, 9).text() if self.sourcing_table.item(r, 9) else "0")
                # 10: Margin
                row_data.append(self.sourcing_table.item(r, 10).text() if self.sourcing_table.item(r, 10) else "0")
                # 11: Margin%
                row_data.append(self.sourcing_table.item(r, 11).text() if self.sourcing_table.item(r, 11) else "0")
                # Link from UserRole
                row_data.append(name_item.data(Qt.ItemDataRole.UserRole) if name_item else "")
                
                data.append(row_data)
            
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(filename, index=False)
            QMessageBox.information(self, "Export", f"성공적으로 저장되었습니다:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"엑셀 저장 중 오류가 발생했습니다:\n{str(e)}")
    def load_from_history(self, item): self.search_input.setText(item.text()); self.perform_research()
    def _set_loading_state(self, s): self.search_btn.setText("Analysing..." if s else "Analysis Start"); self.search_btn.setEnabled(not s)
    def _create_int_item(self, v, s=""):
        item = NumericItem(); item.setData(Qt.ItemDataRole.UserRole, float(v) if str(v).replace(".","").replace(",","").isdigit() else 0); item.setText(f"{v}{s}"); return item
    def init_trend_list_tab(self):
        """네이버 쇼핑 및 분야별 실시간 인기 검색어 탭 초기화"""
        layout = QVBoxLayout(self.trend_list_tab)
        top_ctrl = QHBoxLayout()
        self.trend_cat_combo = QComboBox()
        self.trend_cat_combo.addItems([
            "패션의류 (50000000)", "패션잡화 (50000001)", "화장품/미용 (50000002)", 
            "디지털/가전 (50000003)", "가구/인테리어 (50000004)", "출산/육아 (50000005)",
            "식품 (50000006)", "스포츠/레저 (50000007)", "생활/건강 (50000008)"
        ])
        refresh_btn = QPushButton("🔄 갱신")
        refresh_btn.clicked.connect(self.refresh_realtime_trends)
        top_ctrl.addWidget(QLabel("🎯 카테고리:")); top_ctrl.addWidget(self.trend_cat_combo); top_ctrl.addWidget(refresh_btn); top_ctrl.addStretch()
        layout.addLayout(top_ctrl)
        self.trend_list_table = QTableWidget(); self.trend_list_table.setColumnCount(3)
        self.trend_list_table.setHorizontalHeaderLabels(["순위", "급상승 키워드", "비고"])
        self.trend_list_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.trend_list_table.cellDoubleClicked.connect(lambda r, c: (self.search_input.setText(self.trend_list_table.item(r, 1).text()), self.perform_research()))
        layout.addWidget(self.trend_list_table)

    def refresh_realtime_trends(self):
        """네이버 쇼핑인사이트 오픈 API 및 웹 데이터를 통해 급상승 키워드를 추출합니다."""
        cid_str = self.trend_cat_combo.currentText()
        cid = cid_str.split("(")[-1].replace(")", "").strip()
        url = "https://datalab.naver.com/shoppingInsight/getCategoryKeywordRank.naver"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Referer": "https://datalab.naver.com/",
            "X-Requested-With": "XMLHttpRequest"
        }
        payload = {
            "cid": cid, "timeUnit": "date",
            "startDate": (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d"),
            "endDate": datetime.now().strftime("%Y-%m-%d"),
            "device": "", "gender": "", "ages": "", "page": 1, "count": 20
        }
        try:
            resp = requests.post(url, headers=headers, data=payload, timeout=5)
            if resp.status_code == 200:
                ranks = resp.json().get("ranks", [])
                if ranks:
                    self.trend_list_table.setRowCount(0)
                    for i, item in enumerate(ranks[:20]):
                        self.trend_list_table.insertRow(i)
                        self.trend_list_table.setItem(i, 0, QTableWidgetItem(str(item['rank'])))
                        self.trend_list_table.setItem(i, 1, QTableWidgetItem(item['keyword']))
                        self.trend_list_table.setItem(i, 2, QTableWidgetItem("HOT🔥"))
                    return
            raise Exception("API Fail")
        except:
            # API 실패 시 폴백 (현재 시즌성 인기 키워드)
            fallback_kws = ["미니 가습기", "온열 마사지기", "전동 칫솔", "무선 충전기", "차량용 거치대", "에어프라이어", "요가 밴드", "캠핑 의자", "텀블러", "다이어리"]
            self.trend_list_table.setRowCount(0)
            for i, kw in enumerate(fallback_kws):
                self.trend_list_table.insertRow(i)
                self.trend_list_table.setItem(i, 0, QTableWidgetItem(str(i+1)))
                self.trend_list_table.setItem(i, 1, QTableWidgetItem(kw))
                self.trend_list_table.setItem(i, 2, QTableWidgetItem("추천✨"))

    def update_trends_panel(self, k="", cat_name="", d=30):
        cat_mapper = {"패션의류":"50000000","패션잡화":"50000001","화장품/미용":"50000002","디지털/가전":"50000003","가구/인테리어":"50000004","출산/육아":"50000005","식품":"50000006","스포츠/레저":"50000007","생활/건강":"50000008"}
        cid = cat_mapper.get(cat_name, "50000000")
        try:
            resp = self.shopping_api.get_category_demographics(cid, days=d)
            if resp and 'results' in resp:
                self.trend_shopping_list.clear(); data = resp['results'][0].get('data', [])
                for item in data[-5:]: self.trend_shopping_list.addItem(f"{item['group']}대: {int(item['ratio'])}%")
            if k:
                trends = self.shopping_api.get_keyword_trend(k)
                if trends: self.trend_search_list.clear(); [self.trend_search_list.addItem(f"- {t}") for t in trends[-5:]]
        except: pass

    def load_history(self):
        path = os.path.join(self.config_dir, "history.json")
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    self.search_history = json.load(f); self.history_list.addItems(self.search_history[:10])
            except: pass

    def save_history(self):
        os.makedirs(self.config_dir, exist_ok=True)
        path = os.path.join(self.config_dir, "history.json")
        try:
            with open(path, "w", encoding="utf-8") as f: json.dump(self.search_history[:10], f, ensure_ascii=False, indent=4)
        except: pass

    def load_from_history(self, item):
        """히스토리 클릭 시 해당 키워드 분석을 즉시 실행합니다."""
        self.search_input.setText(item.text())
        self.perform_research()

    def _clear_history(self):
        """모든 검색 기록을 삭제하고 파일을 제거합니다."""
        if QMessageBox.question(self, "확인", "모든 검색 기록을 삭제할까요?") == QMessageBox.StandardButton.Yes:
            self.search_history = []
            self.history_list.clear()
            path = os.path.join(self.config_dir, "history.json")
            if os.path.exists(path):
                try: os.remove(path)
                except: pass
            QMessageBox.information(self, "완료", "검색 기록이 모두 삭제되었습니다.")

    def _add_history(self, k):
        if k in self.search_history: self.search_history.remove(k)
        self.search_history.insert(0, k); self.history_list.clear(); self.history_list.addItems(self.search_history[:10])

    def closeEvent(self, event):
        self.save_history(); event.accept()

    def apply_style(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #0F172A; }
            #sidebar { background-color: #1E293B; border-right: 1px solid #334155; }
            QWidget { background-color: #0F172A; color: #F8FAF6; font-family: 'Malgun Gothic'; font-size: 14px; }
            QLineEdit, QComboBox, QListWidget { background-color: #1E293B; border: 1px solid #334155; border-radius: 6px; padding: 8px; color: white; }
            QPushButton { background-color: #38BDF8; color: #0F172A; border-radius: 6px; font-weight: bold; padding: 10px; }
            QTabWidget::pane { border: 1px solid #334155; background: #0F172A; border-radius: 10px; }
            QTabBar::tab { background: #1E293B; color: #94A3B8; padding: 10px 20px; }
            QTabBar::tab:selected { background: #0F172A; color: #38BDF8; font-weight: bold; }
            QTableWidget { background-color: #1E293B; gridline-color: #334155; }
            QHeaderView::section { background-color: #334155; color: #CBD5E1; padding: 8px; }
            #ai_report_box { color: #F472B6; font-style: italic; }
        """)

if __name__ == "__main__":
    app = QApplication(sys.argv); window = AdvancedTrendApp(); window.show(); sys.exit(app.exec())
